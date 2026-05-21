import openpyxl
from pathlib import Path
import sys

XLSX_NEW = r"c:\Users\User\Downloads\new_pipeline\new_pipeline\categorization_new\Claim-to-Reference Mapping.pre-2026-05-20 (1).xlsx"
OUT = r"c:\Users\User\Downloads\new_pipeline\new_pipeline\scratch\xlsx_dump.txt"

def dump_sheet(ws, max_rows=9999):
    rows_out = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= max_rows:
            rows_out.append(f"  ... (truncated at {max_rows} rows)")
            break
        if all(c is None or str(c).strip() == "" for c in row):
            continue
        cells = [str(c).strip() if c is not None else "" for c in row]
        rows_out.append("  | " + " | ".join(cells) + " |")
    return "\n".join(rows_out)

wb = openpyxl.load_workbook(XLSX_NEW)

lines = []
lines.append("=" * 80)
lines.append(f"FILE: {Path(XLSX_NEW).name}")
lines.append(f"SHEETS ({len(wb.sheetnames)}): {wb.sheetnames}")
lines.append("=" * 80)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    lines.append(f"\n{'=' * 80}")
    lines.append(f"SHEET: '{sheet_name}'  |  Dimensions: {ws.dimensions}  |  Rows: {ws.max_row}  |  Cols: {ws.max_column}")
    lines.append("=" * 80)
    lines.append(dump_sheet(ws, max_rows=600))

lines.append("\n\nDONE.")

Path(OUT).write_text("\n".join(lines), encoding="utf-8")
print(f"Written to {OUT}")
