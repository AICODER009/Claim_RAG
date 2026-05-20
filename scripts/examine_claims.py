"""Show claims file structure correctly."""
import sys
sys.path.insert(0, r"D:\pip_libs")
sys.stdout.reconfigure(encoding="utf-8", errors="replace")
import openpyxl
from collections import Counter
from pathlib import Path

fp = Path(r"D:\revisto_evidence_aligned_clean\new_pipeline\claims\ALL_CLAIMS_COMBINED_categorized_v5.xlsx")
wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]
rows = list(ws.iter_rows(values_only=True))
headers = rows[0]

print("COLUMNS:")
for i, h in enumerate(headers):
    print(f"  [{i}] {h}")

# Category is col 10, Claim Classification is col 12
CAT_COL = 10
CT_COL = 12
CLAIM_COL = 4

cats = Counter()
for r in rows[1:]:
    val = str(r[CAT_COL]).strip() if r[CAT_COL] else "(blank)"
    cats[val] += 1

print(f"\nCATEGORY BREAKDOWN (col 'Category'):")
for c, n in cats.most_common():
    print(f"  {c:20s}: {n}")
print(f"  TOTAL: {sum(cats.values())}")

# CT-ID for originals
ct_counts = Counter()
originals_with_ct = 0
originals_no_ct = 0
for r in rows[1:]:
    cat = str(r[CAT_COL]).strip() if r[CAT_COL] else ""
    if cat == "Original":
        ct = str(r[CT_COL]).strip() if r[CT_COL] and str(r[CT_COL]).strip() not in ("None","") else None
        if ct:
            ct_counts[ct] += 1
            originals_with_ct += 1
        else:
            originals_no_ct += 1

print(f"\nORIGINAL CLAIMS: {originals_with_ct + originals_no_ct}")
print(f"  With CT-ID:    {originals_with_ct}")
print(f"  Without CT-ID: {originals_no_ct}")
print(f"\nCT-ID distribution (top 20):")
for ct, n in ct_counts.most_common(20):
    print(f"  {ct:12s}: {n}")

# Sample originals
print(f"\nSAMPLE ORIGINAL CLAIMS (first 8):")
count = 0
for r in rows[1:]:
    if str(r[CAT_COL]).strip() == "Original" and count < 8:
        claim = str(r[CLAIM_COL])[:130] if r[CLAIM_COL] else ""
        ct = str(r[CT_COL]).strip() if r[CT_COL] else "?"
        refs = str(r[6])[:60] if r[6] else ""
        print(f"  [{ct:8s}] {claim}")
        if refs and refs != "None":
            print(f"            refs: {refs}")
        count += 1

wb.close()
